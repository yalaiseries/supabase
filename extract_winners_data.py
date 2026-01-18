"""
Extract winners data from source documents and generate JSON files for upload.
"""
import json
from pathlib import Path
import sys

try:
    from docx import Document
    import openpyxl
except ImportError:
    print("Installing required packages...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "python-docx", "openpyxl"])
    from docx import Document
    import openpyxl

def extract_2024_data(docx_path):
    """Extract 2024 Programme data from DOCX file."""
    doc = Document(docx_path)
    
    # Extract team data from tables
    teams = []
    
    for table in doc.tables:
        # Skip if not enough rows
        if len(table.rows) < 2:
            continue
            
        # Check if this is a team data table (has columns like Topic, Representative, etc.)
        header_row = table.rows[0]
        headers = [cell.text.strip().lower() for cell in header_row.cells]
        
        # Look for key headers
        if not any(keyword in ' '.join(headers) for keyword in ['topic', 'team', 'representative']):
            continue
        
        # Process each data row
        for row in table.rows[1:]:
            cells = [cell.text.strip() for cell in row.cells]
            if len(cells) < 2 or not cells[0]:  # Skip empty rows
                continue
            
            # Build team entry
            team_data = {}
            for i, header in enumerate(headers):
                if i < len(cells):
                    team_data[header] = cells[i]
            
            if team_data:
                teams.append(team_data)
    
    # Convert to standard format
    use_cases = []
    for team in teams:
        use_case = {
            "title": team.get("topic", team.get("title", "")),
            "team": team.get("team", ""),
            "award": team.get("prize", team.get("award", "")),
            "summary": team.get("about", team.get("summary", "")),
        }
        
        # Add people information
        people = {}
        if team.get("representative speaker"):
            people["representativeSpeaker"] = team["representative speaker"]
        if team.get("designation"):
            people["designation"] = team["designation"]
        if team.get("lead"):
            people["lead"] = team["lead"]
        
        if people:
            use_case["people"] = people
        
        # Add links
        links = []
        if team.get("slide"):
            links.append({"label": "Slides", "url": team["slide"]})
        if links:
            use_case["links"] = links
        
        use_cases.append(use_case)
    
    return {
        "year": 2024,
        "categories": [
            {
                "category": "AI Programme Winners",
                "useCases": use_cases
            }
        ]
    }

def extract_2025_data(xlsx_path):
    """Extract 2025 Hackathon data from XLSX file."""
    wb = openpyxl.load_workbook(xlsx_path)
    
    categories = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        # Process data rows
        use_cases = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            # Map row data to headers
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_data[headers[i].lower()] = str(value).strip() if value else ""
            
            use_case = {
                "title": row_data.get("topic", row_data.get("title", "")),
                "team": row_data.get("team", ""),
                "award": row_data.get("prize", row_data.get("award", "")),
                "summary": row_data.get("one sentence summary", row_data.get("summary", "")),
            }
            
            # Add people
            people = {}
            if row_data.get("representative speaker"):
                people["representativeSpeaker"] = row_data["representative speaker"]
            if row_data.get("designation"):
                people["designation"] = row_data["designation"]
            if row_data.get("lead"):
                people["lead"] = row_data["lead"]
            
            if people:
                use_case["people"] = people
            
            # Add links
            links = []
            if row_data.get("slide"):
                links.append({"label": "Slides", "url": row_data["slide"]})
            if links:
                use_case["links"] = links
            
            use_cases.append(use_case)
        
        if use_cases:
            # Use sheet name as category
            category_name = sheet_name if sheet_name != "Sheet1" else "Top Winners"
            categories[category_name] = use_cases
    
    # Build final structure
    cats = []
    for cat_name, use_cases in categories.items():
        cats.append({
            "category": cat_name,
            "useCases": use_cases
        })
    
    return {
        "year": 2025,
        "categories": cats
    }

def main():
    data_dir = Path(__file__).parent / "data"
    
    # Extract 2024 data
    print("Extracting 2024 data from DOCX...")
    docx_path = data_dir / "2024_AI_Programme.docx"
    data_2024 = extract_2024_data(docx_path)
    
    output_2024 = data_dir / "winners-2024.json"
    with open(output_2024, "w", encoding="utf-8") as f:
        json.dump(data_2024, f, indent=2, ensure_ascii=False)
    print(f"✓ Saved to {output_2024}")
    print(f"  - Teams: {len(data_2024['categories'][0]['useCases'])}")
    
    # Extract 2025 data
    print("\nExtracting 2025 data from XLSX...")
    xlsx_path = data_dir / "2025_AI_Hackathon_Open_Sharing.xlsx"
    data_2025 = extract_2025_data(xlsx_path)
    
    output_2025 = data_dir / "winners-2025.json"
    with open(output_2025, "w", encoding="utf-8") as f:
        json.dump(data_2025, f, indent=2, ensure_ascii=False)
    print(f"✓ Saved to {output_2025}")
    for cat in data_2025['categories']:
        print(f"  - {cat['category']}: {len(cat['useCases'])} entries")
    
    print("\n✅ Data extraction complete!")
    print(f"\nNext steps:")
    print(f"1. Review the generated JSON files")
    print(f"2. Upload to database using winners-admin endpoint")

if __name__ == "__main__":
    main()
