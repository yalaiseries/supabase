"""
Fixed version - properly extract 2025 data from XLSX
"""
import openpyxl
import json

wb = openpyxl.load_workbook(r"C:\2026_AI_Collaboration\aiseries\data\2025_AI_Hackathon_Open_Sharing.xlsx")

categories = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Get headers from first row
    headers = []
    for cell in ws[1]:
        header = str(cell.value).strip() if cell.value else ""
        headers.append(header.lower())
    
    print(f"\nSheet: {sheet_name}")
    print(f"Headers: {headers}")
    
    # Find key column indices
    topic_idx = next((i for i, h in enumerate(headers) if 'topic' in h), None)
    team_idx = next((i for i, h in enumerate(headers) if h == 'team'), None)
    prize_idx = next((i for i, h in enumerate(headers) if 'prize' in h), None)
    summary_idx = next((i for i, h in enumerate(headers) if 'summary' in h or 'sentence' in h), None)
    rep_idx = next((i for i, h in enumerate(headers) if 'representative' in h), None)
    
    print(f"Key columns - Topic: {topic_idx}, Team: {team_idx}, Prize: {prize_idx}, Summary: {summary_idx}")
    
    use_cases = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip if no topic
        if topic_idx is None or not row[topic_idx]:
            continue
        
        topic = str(row[topic_idx]).strip()
        
        # Skip header-like rows
        if topic.lower() in ['topic', 'representative speaker', '']:
            continue
        
        use_case = {
            "title": topic,
        }
        
        if team_idx is not None and row[team_idx]:
            use_case["team"] = str(row[team_idx]).strip()
        
        if prize_idx is not None and row[prize_idx]:
            use_case["award"] = str(row[prize_idx]).strip()
        
        if summary_idx is not None and row[summary_idx]:
            use_case["summary"] = str(row[summary_idx]).strip()
        
        # Add people if available
        if rep_idx is not None and row[rep_idx]:
            use_case["people"] = {
                "representativeSpeaker": str(row[rep_idx]).strip()
            }
        
        use_cases.append(use_case)
        print(f"  Row {row_idx}: {topic[:50]}")
    
    if use_cases:
        categories.append({
            "category": sheet_name if sheet_name != "Sheet1" else "Top Winners",
            "useCases": use_cases
        })
        print(f"  ✓ Added {len(use_cases)} entries")

output = {
    "year": 2025,
    "categories": categories
}

with open(r"C:\2026_AI_Collaboration\aiseries\data\winners-2025.json", "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"\n✓ Saved 2025 data with {sum(len(cat['useCases']) for cat in categories)} total entries across {len(categories)} categories")
